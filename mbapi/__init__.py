# -*- coding: utf-8 -*-
import re
import io
import logging
import json
import time
import uuid
from types import MethodType
from collections import namedtuple, deque

import requests
import pandas as pd
from retry import retry
from lxml import html
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from .product import ProductApi
from .constant import (
    COUNTRY_CODE_MAP,
    CURRENCY_ID_MAP,
    EXPORT_FIELD_MAP,
    MB_BASE_URL,
    AAMZ_BASE_URL,
    VOTOBO_BASE_URL,
    AAMZ_API,
)
from .exceptions import (
    MBApiError,
    MBApiRequestError,
    ProductNoExistError,
    ProductMultiError,
    OrderNotExistError,
    LoginError,
    NotMergedOrderError,
    MBApiBizError,
)
from .config import (
    STOCK_WAREHOUSE_ID, STOCK_GRID_ID, ORDER_UPLOAD_TEMPLATE_ID_MAP, ORDER_DOWNLOAD_TEMPLATE_ID_MAP
)
from .product import ProductSearchOperate, Product


logger = logging.getLogger(__name__)

API_MAP = {
    'login': '%s/index.php?mod=main.doLogin' % MB_BASE_URL,
    'get_product_info': '%s/index.php?mod=stock.getStockList' % AAMZ_BASE_URL,
    'upload_virtual_sku': '%s/index.php?mod=uploadfile.doUploadFileForStock' % AAMZ_BASE_URL,
    'get_shipping_fee': '%s/index.php?mod=order.freightcalculated' % MB_BASE_URL,
    'index': '%s/' % MB_BASE_URL,
    'upload_image': 'https://publish.mabangerp.com/index.php?m=image&a=doUpload',
    'create_order': '%s/index.php?mod=order.doAddOrder' % AAMZ_BASE_URL,
    'get_order_op_log': '%s/index.php?mod=order.getOrderLog' % AAMZ_BASE_URL,
    'search_order': '%s/index.php?mod=order.orderSearch' % MB_BASE_URL,
    # 智能合并订单
    'auto_merge_order': '%s/index.php?mod=order.doAutomationMergeOrder' % MB_BASE_URL,
    # 订单批量上次
    'upload_order_xlsx': '%s/index.php?mod=order.doImportByTemplateData' % MB_BASE_URL,
    # 下载订单表格
    'download_order_xlsx': '%s/index.php?mod=order.doExportByTemplateData' % AAMZ_BASE_URL,
    # 获取订单上次任务状态
    'get_upload_order_status': '%s/index.php?mod=importSystem.getRunningResult' % AAMZ_BASE_URL,
    # 立即执行物流匹配脚本
    'start_ship_match_script': '%s/index.php?mod=ordera.addLogisticsSearch' % MB_BASE_URL,
    # 获取相关订单
    'related_order': f'{AAMZ_BASE_URL}/index.php?mod=order.findrelevantinfo',
    # 通过订单编码获取订单的马帮内部id
    'get_order_info': f'{AAMZ_BASE_URL}/index.php?mod=order.getOrderDeclarationInfo',
    # votobo
    ## votobo_login
    'votobo_login': f'{VOTOBO_BASE_URL}/api/index.php?mod=vmain.mbLogin',
    ## 用于检测votobo登录态
    'votobo_check_login': f'{VOTOBO_BASE_URL}/api/index.php?mod=messageNotice.messageList&type=1',
    ## votobo api
    'votobo_api': f'{VOTOBO_BASE_URL}/api/index.php',
    }


ORDER_OP_TYPE_MAP = {
    '合并订单': '合并订单',
    }


PLATFORM_ID_MAP = {
    'other': 99
    }


ShippingInfo = namedtuple('shipping_info', 'order_id shipping_service tracking_no')


class MBApi(ProductApi):
    @retry((MBApiRequestError, LoginError), 3, 1)
    def request(self, url, method, **kw):
        logger.info(f'url={url}, method={method}, kw={kw}')
        headers = {'X-Requested-With': 'XMLHttpRequest'}
        headers.update(kw.pop('headers', {}))
        try:
            r = getattr(self.r_session, method)(url, headers=headers, **kw)
        except requests.exceptions.RequestException as e:
            raise MBApiRequestError('mb无法访问', e)
        logger.info("mb返回: %s", r.text)
        if r.status_code != 200:
            raise MBApiRequestError('请求mb接口出错, 返回状态码为: %s', r.status_code)
        try:
            ret_data = r.json()
        except json.JSONDecodeError as e:
            raise MBApiRequestError('返回非json数据: %s', r.text)
        if not ret_data['success']:
            if "登录信息已超时" in ret_data["message"]:
                self.login()
                raise MBApiRequestError("登录信息已超时，已经重新登录，请重试")
            raise MBApiBizError('请求mb接口出错, 返回数据为: %s', ret_data)
        if ret_data.get("errorMessage"):
            raise MBApiBizError("调用mb接口成功，但出现错误: %s" % ret_data["errorMessage"])
        return ret_data

    def check_login(self):
        aamz_text = self._r_session.get(AAMZ_API).text
        mb_text = self._r_session.get(API_MAP['index']).text
        votobo_json = self._r_session.get(API_MAP['votobo_check_login']).json()
        # 登录标记
        login_flag = '企业编号'
        if login_flag in mb_text and login_flag in aamz_text and votobo_json['success']:
            logger.info('%s 登陆状态正常', self.user)
            self.login_error_times = 0
        else:
            self.login_error_times += 1
            if self.login_error_times < 3:
                return self.login()
            raise LoginError("MB登录失败")

    def login(self):
        # TODO: 处理login和check_login循环调用的问题
        logger.info('登陆mb: %s' % self.user)
        login_api = API_MAP['login']
        data = {'username': self.user, 'password': self.passwd}
        r = self._r_session.post(login_api, data=data)
        r_json = r.json()
        logger.info('登录返回信息: %s', r_json)
        if not r_json['success']:
            raise LoginError('MB登录失败')

        c_mkey = r.cookies["MABANG_ERP_PRO_MEMBERINFO_LOGIN_COOKIE"]

        # 为AAMZ_API注册好Cookie
        aamz_params = {
            'mod': 'stock.list',
            'searchStatus': 3,
            'cMKey': c_mkey,
            'lang': 'cn',
            }
        resp = self._r_session.get(AAMZ_API, params=aamz_params)
        logger.info('登录AAMZ返回信息: %s', resp.text[:150])

        votobo_params = {
            "mod": "vmain.mbLogin",
            "mbkey": f"md_MABANG_ERP_PRIVATE_LOGIN_{self.business_number}_{self.user_id}_M0010806",
            "private_mabang": "",
        }
        resp = self._r_session.get(API_MAP['votobo_login'], params=votobo_params)
        logger.info('登录votobo返回信息: %s', resp.json())
        self.check_login()

    def get_shipping_fee(self, weight, country='US'):
        '''获取邮费, 只支持e邮宝'''
        api = API_MAP['get_shipping_fee']
        data = {
            'countryCode': 'US',
            'orderweiht': weight
            }
        r = self.r_session.post(api, data=data)
        html_str = r.json()['message']
        tree = html.fromstring(html_str)
        shipping_infos = tree.xpath('//body//text()')
        for info in shipping_infos:
            if '线下E邮宝' in info:
                return float(info.rsplit(' ', 1)[-1])
        else:
            raise MBApiError('获取物流价格失败')

    def upload_virtual_sku_from_file(self, xlsx_name):
        '''上传虚拟SKU'''
        api = API_MAP['upload_virtual_sku']
        files = {'templetfile': open(xlsx_name, 'rb')}
        data = {
            "UpLoadFileType": "addVirtualSKU",
            "stockVirtualType": 1
        }
        ret_data = self.request(api, "post", data=data, files=files)
        logger.info('上传sku: %s, 返回数据:%s' % (xlsx_name, ret_data))
        return ret_data

    def upload_virtual_sku(self, mb_sku_map_list):
        '''上传sku对
        :param mb_sku_map_list: (mb_sku, vir_sku)...
        '''
        logger.info('上传sku: %s' % mb_sku_map_list)
        api = API_MAP['upload_virtual_sku']
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '*库存sku编号'
        ws['B1'] = '*虚拟sku1'
        for num, (mb_sku, vir_sku) in enumerate(mb_sku_map_list, 2):
            ws['A%s' % num] = mb_sku
            ws['B%s' % num] = vir_sku
        fp = io.BytesIO(save_virtual_workbook(wb))
        fp.name = 'test.xlsx'
        files = {'templetfile': fp}
        data = {
            "UpLoadFileType": "addVirtualSKU",
            "stockVirtualType": 1
        }
        ret_data = self.request(api, "post", data=data, files=files)
        logger.info('上传sku: %s, 返回数据:%s' % (mb_sku_map_list, ret_data))
        return ret_data

    def exist_virtual_sku(self, vir_sku):
        try:
            if self.get_product_info_from_virtual_sku(vir_sku):
                return True
            else:
                return False
        except ProductNoExistError:
            return False
        except ProductMultiError:
            return True

    def get_newest_virtual_sku(self, vir_sku):
        '''判断该类型的虚拟sku的最小未使用的虚拟sku
        param vir_sku: 根据这个vir_sku返回对应的最小未使用的虚拟sku
        '''
        def next_sku(sku):
            sku_re = re.compile(r'([A-Za-z]+)(\d+)')
            prefix, num = sku_re.match(sku).groups()
            new_num = int(num) + 1
            return '%s%0{}d'.format(len(num)) % (prefix, new_num)

        while True:
            logger.debug('测试vir_sku [%s] 是否存在' % vir_sku)
            if not self.exist_virtual_sku(vir_sku):
                return vir_sku
            vir_sku = next_sku(vir_sku)

    def upload_image(self, img_fp):
        '''上传图片
        :param img_fp: 图片路径或者图片file object
        :return: 上传后图片url
        '''
        if isinstance(img_fp, str):
            img_f = open(img_fp, 'rb')
        else:
            img_f = img_fp
        api = API_MAP['upload_image']
        files = {'UpLoadFile': ('test.jpg', img_f, 'image/jpeg')}
        data = {'postName': 'UpLoadFile'}
        r = requests.post(api, data=data, files=files)
        logger.debug('%s: %s' % (r.status_code, r.text))
        return r.json()['imageUrl']

    def get_profit_info(self, shopname, start_date, end_date):
        '''通过报表查询各商品利润信息'''

    def create_order(self, vir_sku, quantity, price, **params):
        '''创建订单
        :param vir_sku: 虚拟sku
        :param quantity: 商品数量
        :param price: 商品单价
        其他参数说明见接口参数
        '''
        # TODO: 未完成的接口
        api = API_MAP['create_order']
        common_data = dict(
            platformOrderId=platformOrderId,
            salesRecordNumber=salesRecordNumber,
            shopId=shopId,
            platformId=platformId,
            currencyId=currencyId,
            itemTotal=itemTotal,
            shippingFee=shippingFee,
            platformFee=platformFee,
            insuranceFee='',
            paypalFee='',
            paypalEmail='',
            paidTime=paidTime,
            buyerUserId=buyerUserId,
            buyerName=buyerName,
            phone1=phone1,
            phone2='',
            email=email,
            countryCodeN=COUNTRY_CODE_MAP[countryCodeid],
            countryCodeid=countryCodeid,
            province=province,
            city=city,
            postCode=postCode,
            doorcode='',
            street1=street1,
            VendorID='',
            abnnumber='',
            shippingService='',
            myLogisticsChannelId=0,
            trackNumber='',
            trackNumber2='',
            )
        product = self.get_product_info_from_virtual_sku(vir_sku)
        product_params = {
            'stockSkuA[]': self.sku,
            }

    def get_order_info(self, order_id):
        '''获取订单的信息'''
        api = API_MAP['get_order_info']
        data = {
            'tableBase': 1,
            'platformOrderId': order_id,
            'myLogisticsChannelId': '',
            # 该参数可不传，该参数未马帮内部订单id
            # 'orderId': ''
            }
        ret_data = self.request(api, 'post', data=data)
        pos_html = ret_data['posHtml']
        raw_json_data = re.search(r'(?<=>){.*}(?=<)', pos_html).group()
        return json.loads(raw_json_data)

    def get_mb_order_id(self, order_id):
        '''获取订单的马帮内部id
        :param order_id: 订单编号
        :return: 马帮内部订单id
        '''
        api = AAMZ_API
        params = {
            'mod': 'order.detail',
            'platformOrderId': order_id,
            'orderStatus': 2,
            'orderTable': 2,
            'tableBase': 2,
            'lang': 'cn',
            }
        html_text = self.r_session.get(api, params=params).text
        mb_order_id = re.search(r'(?<=&orderId=)\d+', html_text).group()
        return int(mb_order_id)

    def get_order_op_log(self, mb_order_id):
        '''获取订单的操作日志
        :return: 操作日志字典列表，默认排序为mb返回的排序，即时间倒序.
            格式: [{'操作属性': x, '描述': x, '操作员': x, '操作时间': x, '其他信息': {}}]
        '''
        def handle_tr(tr):
            data = {}
            op_type = tr.xpath('./td[1]/text()')[0]
            data['op_type'] = op_type
            data['detail'] = ''.join(tr.xpath('./td[2]//text()'))
            data['operator'] = tr.xpath('./td[3]/text()')[0]
            data['op_time'] = tr.xpath('./td[4]/text()')[0]
            data['ext'] = {}
            if op_type == ORDER_OP_TYPE_MAP['合并订单']:
                # 是否未合并订单的主订单号
                if '合并到订单' in data['detail']:
                    data['ext']['order_id'] = tr.xpath('./td[2]/a/text()')[0]
                    data['ext']['is_main'] = False
                else:
                    data['ext']['order_id'] = tr.xpath('./td[2]/a/text()')
                    data['ext']['is_main'] = True
            return data

        api = API_MAP['get_order_op_log']
        data = {
            'htmltype': 'tr',
            'orderId': mb_order_id,
            'page': '',
            'rowsPerPage': ''
            }
        ret_data = self.request(api, 'post', data=data)
        html_text = ret_data['message']
        tree = html.fromstring(html_text)
        return [
            handle_tr(tr)
            for tr in tree.xpath('//tr')
            ]

    def get_order(self, order_id: str):
        '''搜索订单'''
        api = API_MAP['search_order']
        data = {
            'OrderSearch.fuzzySearchKey': 'Order.platformOrderId',
            'OrderSearchFuSKey': 'a.platformOrderId',
            'daysOperator': '=',
            'OrderSearch.fuzzySearchValue': order_id,
            'orderPageKey': uuid.uuid1().hex,
            'a': 'orderalllist',
            'post_tableBase': 1,
            }
        ret_data = self.request(api, 'post', data=data)
        order_list = ret_data['orderDataList']
        if not order_list:
            raise MBApiError(f'{order_id} 查无该订单')
        if len(order_list) > 1:
            raise MBApiError(f'{order_id} 查询出了多个订单')
        return order_list[0]

    def get_order_logistics_info(self, order_id: str):
        '''获取订单物流信息'''
        order = self.get_order(order_id)
        logistics_html = order['cansend1logisticsHtml']
        tree = html.fromstring(logistics_html)
        ship_serv = tree.xpath('./p[1]/text()')[0].strip()
        tracking_no = order['trackNumber']
        return {'ship_serv': ship_serv, 'tracking_no': tracking_no}

    def get_order_by_ids(self, order_ids: list):
        '''获取搜索多个订单信息'''
        api = API_MAP['search_order']
        data = {
            'platformTracknumberSearchInput': 'platformOrderId',
            'platformTracknumberSearchtextarea': '\n'.join(order_ids)
            }
        return self.request(api, 'post', data=data)

    def get_order_shipping_info_by_ids(self, order_ids: list) -> list:
        '''获取物流信息
        :return: 返回格式[{order_id: x, shipping_service: x, tracking_no: x}]和不存在的订单id
        '''
        def get_merge_order_shipping_info(order_id):
            op_log = self.get_order_op_log(order_id)['message']
            main_order_id = re.search(r'合并到订单<a[^>]+>(\w+)</a>', op_log).group(1)
            return self.get_order_shipping_info(main_order_id)

        def convert(order_data):
            no_shipping_keyword = [
                'title="物流渠道未选择"',
                'title="无运单号"'
                ]
            order_id = order_data['platformOrderId']
            shipping_info = order_data['cansend1logisticsHtml']
            if order_data['showOrderStatusText'] == '已作废' and '合并订单' in order_data['order_label']:
                return get_merge_order_shipping_info(order_id)

            if any(kw in shipping_info for kw in no_shipping_keyword):
                shipping_service = ''
                tracking_no = ''
            else:
                shipping_service, tracking_no = re.findall(r'(?<=>)[^<]+(?=<)', shipping_info)
            return ShippingInfo(
                order_id=order_id,
                shipping_service=shipping_service,
                tracking_no=tracking_no,
                )

        def get_no_exist_ids(order_list, order_ids):
            '''获取不存在马帮的订单'''
            if len(order_list) != len(order_ids):
                no_exist_ids = list(
                    set(order_ids) - set(order['platformOrderId'] for order in order_list)
                    )
            else:
                no_exist_ids = []
            return no_exist_ids

        ret_data = self.get_order_by_ids(order_ids)
        order_list = ret_data['orderDataList']
        no_exist_ids = get_no_exist_ids(order_list, order_ids)
        shipping_info_list = [convert(order) for order in order_list]
        return shipping_info_list, no_exist_ids

    def get_order_shipping_info(self, order_id):
        '''获取单订单的物流信息'''
        shipping_info_list, no_exist_ids = self.get_order_shipping_info_by_ids([order_id])
        if not shipping_info_list:
            raise OrderNotExistError('订单不存在, 订单id:  %s', order_id)
        return shipping_info_list[0]

    def auto_merge_order(self, shop_id):
        '''智能合并订单'''
        api = API_MAP['auto_merge_order']
        data = [
            ('isWishEpc', ''),
            ('FramePage', ''),
            ('type', 'NEUB'),
            ('platform', ''),
            ('Order.shops[]', shop_id),
            ('mergeCondition[]', 0),
            ('mergeCondition[]', 1),
            ('mergeCondition[]', 10),
            ('mergeCondition[]', 2),
            ('mergeCondition[]', 4),
            ('mergeCondition[]', 5),
            ('buyersAccount', ''),
            ('type', 'FEUB'),
            ('mergeRemark', '同姓名,同客户ID,同邮寄地址,订单重量超过2kg不合并,拆分订单不合并'),
            ('tableBase', 1),
            ('checkOrderSecLog', 1),
            ('remarkflag', 1),
            ('changeprint', ''),
        ]
        return self.request(api, 'post', data=data)

    def _upload_order_xlsx(self, fp, template_id, shop_id):
        '''上传订单文件
        :param fp: 文件对象，必须有name属性; 或文件路径
        '''
        if isinstance(fp, str):
            fp = open(fp, 'rb')
        api = API_MAP['upload_order_xlsx']
        file_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        files = {'templetfile': (fp.name, fp, file_content_type)}
        data = {'templateId': template_id, 'shopId': shop_id}
        resp = self.r_session.post(api, data=data, files=files)
        if '"success":true' not in resp.text:
            raise MBApiError('订单文件上传失败，返回信息为: %s', resp.text)
        return resp.text

    def upload_order_xlsx_for_5miles(self, fp, shop_id):
        '''上传5miles订单'''
        return self._upload_order_xlsx(fp, ORDER_UPLOAD_TEMPLATE_ID_MAP['5miles'], shop_id)

    def export_order(self, order_ids: list, headers: list, template_id: int=0) -> list:
        '''导出订单信息
        :param order_ids: 订单id列表
        :param headers: 导出的字段列表
        :param: template_id: 导出订单由headers确定，目前template_id可不传入
        :return: 返回订单列表
        '''
        assert order_ids
        api = API_MAP['download_order_xlsx']
        data = [
            ('backUrl', ''),
            ('orderIds', '\n'.join(order_ids)),
            ]
        for field in headers:
            data.append(('fieldlabel', EXPORT_FIELD_MAP[field]))
        for field in headers:
            data.extend([
                ('map-name[]', field),
                ('map-uq[]', EXPORT_FIELD_MAP[field]),
                ('map-text[]', '')
            ])
        data.extend([
            ('templateName', ''),
            ('templateId', template_id),
            ('standardVersion', '1'),
            ('orderItemOrderBy', ''),
            ('pageSave', '1'),
            ('tableBase', ''),
            # 合并多品信息
            ('mergeShow', 1),
            ('hbddgyxx', 2),
            ])
        url = self.request(api, 'post', data=data)['gourl']
        content = self.r_session.get(url).content
        df = pd.read_excel(io.BytesIO(content), na_filter=False)
        ret_data = df.values.tolist()
        if len(ret_data) != len(order_ids):
            raise MBApiError(f'导出订单接口错误, 导出前后订单数量[{len(order_ids),len(ret_data)}]不一致')
        return ret_data

    def download_order_xlsx_for_5miles(self, order_ids: list) -> list:
        '''下载5miles订单表格
        注意:
            一般情况下，导打算导出的订单数据等于出的实际订单数量.
            如果存在该批次订单与之前批次订单合并的话，则后者大于前者
        :return: 订单数据列表
        '''
        def convert_shipping_service(name):
            '''转换物流编号
            :param name: 马帮的物流中文名
            '''
            name_map = {
                '国际电商专递': 'sfb2c',
                '顺丰': 'sfb2c',
                'E邮宝': 'china-ems',
                '燕文': 'yanwen',
                '燕邮宝': 'yanwen',
                '递四方': '4px',
                '联邮通': '4px',
            }
            for k, v in name_map.items():
                if k in name:
                    return v
            else:
                raise ValueError('获取不到[%s]对应的5miles代码' % name)

        def unfold(row):
            '''展开所有合并订单'''
            order_id_list = [
                row[order_id_index],
                # 被合并订单会生成多列
                *get_merge_order_ids(row, order_id_index, merge_order_index)
                ]
            sku_list = list(filter(bool, row[sku_index].split(';')))
            quantity_list = list(map(int, filter(bool, str(row[quantity_index]).split(';'))))
            if not (len(order_id_list) == len(sku_list) == len(quantity_list)):
                raise MBApiError(f'{row[order_id_index]}, sku数量和商品数量和订单的长度不一致')
            return order_id_list, sku_list, quantity_list

        def get_merge_order_ids(row, order_id_index, merge_order_start_index):
            '''取出被合并到该订单的列表，即不包含自身订单id
            由于重发订单或者重新激活的订单，订单号会加后缀如"_1",
            但导出时订单号列会去除后缀，并且在合并订单列重复该订单号,
            因此得去除
            '''
            merge_order_ids = list(filter(bool, row[merge_order_index:]))
            try:
                merge_order_ids.remove(row[order_id_index])
            except ValueError:
                pass
            return merge_order_ids

        headers = ['交易编号', '平台SKU', '物流渠道', '商品数量', '货运单号', '交运异常原因', '状态', '被合并订单']
        order_id_index = 0
        sku_index = 1
        ship_serv_index = 2
        quantity_index = 3
        track_no_index = 4
        ship_error_index = 5
        status_index = 6
        # 被合并订单会生成多列
        merge_order_index = 7
        order_list = self.export_order(order_ids, headers)
        # 已处理的订单，对于合并的订单，防止订单重复处理
        # 无物流信息订单
        no_info_rows = []
        # 作废订单
        invalid_rows = []
        ret_data = []
        valid_rows = []
        # 获取被合并的订单id列表
        merged_order_id_set = set()
        for row in order_list:
            merged_order_id_set.update(
                get_merge_order_ids(row, order_id_index, merge_order_index)
                )

        # 初步过滤出，有效订单，被合并作废订单，无物流信息订单
        for row in order_list:
            if row[order_id_index] in merged_order_id_set:
                continue
            # 已作废订单
            if not all([row[sku_index], row[quantity_index], row[status_index]!='已作废']):
                invalid_rows.append(row)
                continue
            # 无物流信息的订单
            if not all([row[ship_serv_index], row[track_no_index]]):
                no_info_rows.append(row)
                continue
            # 有效订单
            valid_rows.append(row)

        # 存在无物流信息的订单
        if no_info_rows:
            log = '\n'.join(
                '[%s]订单没有完整的物流信息: 物流商[%s], 物流编号: [%s], 交运异常原因: [%s]'
                % (row[order_id_index], row[ship_serv_index], row[track_no_index], row[ship_error_index])
                for row in no_info_rows
                )
            raise MBApiError(log)

        # 获取作废的订单的物流信息
        for row in invalid_rows:
            # WARNING: 5miles订单编号与交易编号一致，这里使用的是交易编号
            order_id = row[order_id_index]
            logger.info(f'{order_id} 尝试获取合并订单的主订单号的物流信息')
            # 获取合并订单的主订单编号
            main_order_id = self.get_main_order_id(order_id)
            main_order_row = self.export_order([main_order_id], headers)[0]
            valid_rows.append(main_order_row)

        for row in valid_rows:
            order_id_list, sku_list, quantity_list = unfold(row)
            try:
                miles_ship_serv = convert_shipping_service(row[ship_serv_index])
            except ValueError as e:
                raise MBApiError(
                    f'[{row[order_id_index]}]订单物流[{row[ship_serv_index]}]获取不到对应5mile代码'
                    )
            for order_id, sku, quantity in zip(order_id_list, sku_list, quantity_list):
                ret_data.append((
                    order_id, sku, miles_ship_serv, quantity, row[track_no_index]
                    ))

        export_order_id_set = set(item[0] for item in ret_data)
        # 要导出订单可能带有_1后缀, 需进行处理
        replace_re = re.compile(r'_\d+$')
        order_id_set = set(replace_re.sub('', order, 1) for order in order_ids)
        # 实际导出订单是打算导出订单的超集
        if order_id_set - export_order_id_set:
            raise MBApiError(
                '导出订单前后不一致, 少了订单为[%s], 多了订单为[%s]'
                % (order_id_set-export_order_id_set, export_order_id_set-order_id_set)
                )
        return ret_data

    def get_main_order_id(self, order_id):
        '''获取合并订单的主id'''
        mb_order_id = self.get_mb_order_id(order_id)
        op_log_list = self.get_order_op_log(mb_order_id)
        # 拿到最新的合并订单信息
        for item in op_log_list:
            if item['op_type'] == ORDER_OP_TYPE_MAP['合并订单'] and not item['ext']['is_main']:
                return item['ext']['order_id']
        else:
            raise NotMergedOrderError(f'{order_id}非合并订单')

    def get_order_ext_info(self, mb_order_id: int, info_type='related'):
        '''获取相关订单
        :param mb_order_id: 马帮的订单id
        :param info_type: 信息类型
        :return: 订单信息字典列表
        '''
        # TODO: 该接口暂未完善
        api = API_MAP['related_order']
        data = {
            'orderId': mb_order_id,
            'type': 1,
            'tableBase': 2,
            }
        ret_data = self.request(api, 'post', data=data)
        order_html = ret_data['order_html']
        tree = html.fromstring(order_html)
        ret_data = []
        for tr in tree.xpath('//tr'):
            # url = tr.xpath('./td[1]/a/@href')[0]
            订单编号 = tr.xpath('./td[1]/a/text()')[0]
            状态 = tr.xpath('./td[2]/text()')[0]
            # 订单金额 = tr.xpath('./td[3]/text()')[0]
            # 付款时间 = tr.xpath('./td[4]/p[0]/text()')[0]
            发货时间 = tr.xpath('./td[4]/p[1]/text()')[0]
            ret_data.append({'订单编号': 订单编号})
        return ret_data

    def is_order_uploaded(self, filename):
        '''获取订单是否已成功上传'''
        api = API_MAP['get_upload_order_status']
        return filename in self.r_session.get(api).text

    def get_order_upload_status(self, filename):
        '''获取某订单文件的上传状态'''
        OrderStatus = namedtuple('StatusStr', [
            'finish_time',
            'created_time',
            'classification',
            'filename',
            'total',
            'success_num',
            'fail_num',
            'log_url'
            ])
        api = API_MAP['get_upload_order_status']
        text = self.r_session.get(api).text
        status_str_list = text.split('</tr>')
        for status_str in status_str_list:
            if filename in status_str:
                params = re.findall(r'(?<=>)[^<]+(?=<)', status_str)[:7]
                log_url_match = re.search(r"window.open\('([^']+)'\)", status_str)
                log_url = log_url_match and log_url_match.group(1)
                params[4:7] = list(map(int, params[4:7]))
                return OrderStatus(*params, log_url)

    def start_ship_match_script(self):
        '''立即执行物流匹配脚本'''
        api = API_MAP['start_ship_match_script']
        return self.request(api, 'post', data={'type': 2})

    def get_dev_product_detail(self, dev_product_id: int):
        """获取待开发商品的详情
        :param dev_product_id: 待开发商品的id
        """
        api = API_MAP['votobo_api']
        params = {
            "mod": "productApi.getProductDetail",
            "productId": dev_product_id,
        }
        return self.request(api, 'get', params=params)
